# -*- coding: utf-8 -*-
"""
客户出单统计异常分析 & 分群运营分析
输入: D:/joe-project/客保通数据/客户出单统计 (3).xlsx
输出: D:/joe-project/客保通数据/analysis_output/客户异常清单与分群分析_YYYYMMDD.xlsx
"""
import pandas as pd
import numpy as np
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SRC = r"D:/joe-project/客保通数据/客户出单统计 (3).xlsx"
OUT_DIR = r"D:/joe-project/客保通数据/analysis_output"
TODAY = datetime.now().strftime("%Y%m%d")
OUT = f"{OUT_DIR}/客户异常清单与分群分析_{TODAY}.xlsx"

# ---------- 数据准备 ----------
df = pd.read_excel(SRC, sheet_name="sheet1")
df['期'] = pd.to_datetime(df['签单时间'], format='%Y-%m').dt.to_period('M').astype(str)
g = df.groupby(['期', '客户代码'], as_index=False).agg(保费=('保费量', 'sum'), 单量=('出单量', 'sum'))

PERIODS = {
    'Q1_2026': ('2026-01', '2026-03', '2026年Q1'),
    'Q2_2026': ('2026-04', '2026-06', '2026年Q2'),
    'Q4_2025': ('2025-10', '2025-12', '2025年Q4'),
    'M06_2026': ('2026-06', '2026-06', '2026年6月'),
    'M07_2026': ('2026-07', '2026-07', '2026年7月'),
    'H1_2026': ('2026-01', '2026-06', '2026上半年'),
}


def period_sum(key):
    p1, p2, _ = PERIODS[key]
    m = g[(g['期'] >= p1) & (g['期'] <= p2)]
    return m.groupby('客户代码')[['保费', '单量']].sum().reset_index()


# ---------- 异常表 ----------
def anomaly_table(base_key, cur_key):
    a, b = period_sum(base_key), period_sum(cur_key)
    t = a.merge(b, on='客户代码', how='left', suffixes=('_基', '_对'))
    t = t[t['保费_基'] > 0].copy()
    t = t.fillna({'保费_对': 0, '单量_对': 0})
    t['保费环比%'] = (t['保费_对'] - t['保费_基']) / t['保费_基'] * 100
    t['单量环比%'] = (t['单量_对'] - t['单量_基']) / t['单量_基'] * 100
    t['保费降>30%'] = np.where(t['保费环比%'] < -30, '是', '')
    t['单量降>30%'] = np.where(t['单量环比%'] < -30, '是', '')
    return t


T1 = anomaly_table('Q1_2026', 'Q2_2026')
T2 = anomaly_table('Q4_2025', 'Q2_2026')
T3 = anomaly_table('M06_2026', 'M07_2026')

# 排序：先异常(任一)在前，再按基准保费降序
for t in (T1, T2, T3):
    t['_any'] = (t['保费降>30%'] == '是') | (t['单量降>30%'] == '是')
    t.sort_values(['_any', '保费_基'], ascending=[False, False], inplace=True)
    t.drop(columns='_any', inplace=True)

# ---------- 分群（按2026上半年保费排名） ----------
h1 = period_sum('H1_2026').sort_values('保费', ascending=False).reset_index(drop=True)
h1['排名'] = h1.index + 1

q1 = period_sum('Q1_2026'); q2 = period_sum('Q2_2026')
q4 = period_sum('Q4_2025'); m6 = period_sum('M06_2026'); m7 = period_sum('M07_2026')


def segment_detail(seg):
    s = seg.rename(columns={'保费': 'H1保费', '单量': 'H1单量'}).copy()
    for key, pref in [('Q1_2026', 'Q1'), ('Q2_2026', 'Q2'), ('Q4_2025', '25Q4'),
                      ('M06_2026', '6月'), ('M07_2026', '7月')]:
        d = period_sum(key).rename(columns={'保费': f'{pref}保费', '单量': f'{pref}单量'})
        s = s.merge(d, on='客户代码', how='left')
    s = s.fillna(0)
    s['Q2保费环比%'] = np.where(s['Q1保费'] > 0, (s['Q2保费'] - s['Q1保费']) / s['Q1保费'] * 100, np.nan)
    s['Q2同比25Q4%'] = np.where(s['25Q4保费'] > 0, (s['Q2保费'] - s['25Q4保费']) / s['25Q4保费'] * 100, np.nan)
    s['7月环比6月%'] = np.where(s['6月保费'] > 0, (s['7月保费'] - s['6月保费']) / s['6月保费'] * 100, np.nan)
    s['7月状态'] = np.where(s['7月保费'] == 0, '本月沉默', np.where(s['7月保费'] > s['6月保费'], '增长', '下滑'))
    return s


seg_detail_1 = segment_detail(h1[h1['排名'] <= 150])
seg_detail_2 = segment_detail(h1[(h1['排名'] >= 151) & (h1['排名'] <= 400)])
seg_detail_3 = segment_detail(h1[h1['排名'] >= 401])

# ---------- Excel 样式 ----------
HEADER_FILL = PatternFill('solid', fgColor='1F4E78')
HEADER_FONT = Font(color='FFFFFF', bold=True, size=11)
RED_FILL = PatternFill('solid', fgColor='FFC7CE')
RED_FONT = Font(color='9C0006', bold=True)
LIGHT_RED_ROW = PatternFill('solid', fgColor='FDE9E9')
GREEN_FILL = PatternFill('solid', fgColor='C6EFCE')
GREEN_FONT = Font(color='006100')
thin = Side(style='thin', color='D9D9D9')
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
CENTER = Alignment(horizontal='center', vertical='center')
LEFT = Alignment(horizontal='left', vertical='center')

NUM_FMT = '#,##0.00'
INT_FMT = '#,##0'
PCT_FMT = '0.0'

wb = Workbook()

# 说明页
ws0 = wb.active
ws0.title = '说明'
notes = [
    ['客户出单统计 · 异常客户清单与分群运营分析', ''],
    ['', ''],
    ['数据来源', '客户出单统计 (3).xlsx（sheet1，2025-01 至 2026-07，3,010个客户，32,429条记录）'],
    ['统计口径', '保费/单量按“客户代码 × 月份”汇总；负保费为退保/冲减记录，按净值计入'],
    ['异常规则', '保费环比下降超30% 或 单量环比下降超30% 判定为异常（单元格标红）'],
    ['环比公式', '（对比期 - 基准期）/ 基准期 × 100%；基准期为0时不计入（无法计算）'],
    ['客户名称', '源数据仅有“客户代码”字段，无客户名称，请按客户代码对照CRM系统'],
    ['', ''],
    ['Sheet列表', ''],
    ['异常表1_Q2vsQ1', '2026年Q2（4-6月）对比 2026年Q1（1-3月），基准期为Q1有保费客户（2,136户）'],
    ['异常表2_Q2vs25Q4', '2026年Q2（4-6月）对比 2025年Q4（10-12月），基准期为25Q4有保费客户（2,183户）'],
    ['异常表3_7月vs6月', '2026年7月对比 2026年6月，基准期为6月有保费客户（1,627户）'],
    ['分群汇总', '按2026上半年保费排名：前150 / 151-400 / 401+ 三群核心指标对比'],
    ['前150客户明细', '2026上半年保费排名前150客户逐户经营数据'],
    ['151-400客户明细', '2026上半年保费排名151-400客户逐户经营数据'],
    ['401+客户明细', '2026上半年保费排名401+客户逐户经营数据'],
    ['', ''],
    ['筛选建议', '异常清单可优先关注“基准保费≥1万元且保费降幅>30%”的大客户，见各表“保费_基”列'],
    ['使用提醒', '7月仅单月数据，环比受自然波动影响较大，建议结合Q2趋势综合判断'],
]
for r, row in enumerate(notes, 1):
    ws0.cell(r, 1, row[0])
    ws0.cell(r, 2, row[1])
ws0.cell(1, 1).font = Font(bold=True, size=14, color='1F4E78')
ws0.column_dimensions['A'].width = 28
ws0.column_dimensions['B'].width = 110


def write_anomaly_sheet(ws, t, base_label, cur_label):
    cols = ['客户代码', f'{base_label}保费', f'{base_label}单量', f'{cur_label}保费', f'{cur_label}单量',
            '保费环比%', '单量环比%', '保费降>30%', '单量降>30%']
    ws.append(cols)
    for c in range(1, len(cols) + 1):
        cell = ws.cell(1, c)
        cell.fill, cell.font, cell.alignment, cell.border = HEADER_FILL, HEADER_FONT, CENTER, BORDER
    for _, r in t.iterrows():
        any_anomaly = (r['保费降>30%'] == '是') or (r['单量降>30%'] == '是')
        row = [r['客户代码'], r['保费_基'], r['单量_基'], r['保费_对'], r['单量_对'],
               r['保费环比%'], r['单量环比%'], r['保费降>30%'], r['单量降>30%']]
        ws.append(row)
        i = ws.max_row
        for c in range(1, len(cols) + 1):
            cell = ws.cell(i, c)
            cell.border = BORDER
            if c in (2, 4):
                cell.number_format = NUM_FMT
            elif c in (3, 5):
                cell.number_format = INT_FMT
            elif c in (6, 7):
                cell.number_format = PCT_FMT
        if any_anomaly:
            for c in range(1, len(cols) + 1):
                ws.cell(i, c).fill = LIGHT_RED_ROW
        if r['保费降>30%'] == '是':
            ws.cell(i, 6).fill, ws.cell(i, 6).font = RED_FILL, RED_FONT
        if r['单量降>30%'] == '是':
            ws.cell(i, 7).fill, ws.cell(i, 7).font = RED_FILL, RED_FONT
    widths = [12, 14, 12, 14, 12, 11, 11, 12, 12]
    for c, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = 'A2'


ws1 = wb.create_sheet('异常表1_Q2vsQ1')
write_anomaly_sheet(ws1, T1, '2026Q1', '2026Q2')
ws2 = wb.create_sheet('异常表2_Q2vs25Q4')
write_anomaly_sheet(ws2, T2, '2025Q4', '2026Q2')
ws3 = wb.create_sheet('异常表3_7月vs6月')
write_anomaly_sheet(ws3, T3, '2026年6月', '2026年7月')

# ---------- 分群汇总 ----------
seg_meta = [
    ('前150', seg_detail_1, '高价值核心客户（2026上半年保费排名1-150）'),
    ('151-400', seg_detail_2, '中坚成长客户（排名151-400）'),
    ('401+', seg_detail_3, '长尾基础客户（排名401+，共%d户）' % len(seg_detail_3)),
]
ws4 = wb.create_sheet('分群汇总')


def seg_summary(d):
    n = len(d)
    h1p, h1v = d['H1保费'].sum(), d['H1单量'].sum()
    q2_chg = (d['Q2保费'].sum() - d['Q1保费'].sum()) / d['Q1保费'].sum() * 100 if d['Q1保费'].sum() > 0 else np.nan
    q2_yoy = (d['Q2保费'].sum() - d['25Q4保费'].sum()) / d['25Q4保费'].sum() * 100 if d['25Q4保费'].sum() > 0 else np.nan
    m7_chg = (d['7月保费'].sum() - d['6月保费'].sum()) / d['6月保费'].sum() * 100 if d['6月保费'].sum() > 0 else np.nan
    silent = (d['7月保费'] == 0).sum()
    top10_share = d.nlargest(10, 'H1保费')['H1保费'].sum() / h1p * 100 if h1p > 0 else 0
    decl_q2 = (d['Q2保费环比%'] < -30).sum()
    return {
        '客户数': n, 'H1保费万': round(h1p / 1e4, 1), 'H1单量': int(h1v),
        '户均保费': round(h1p / n), '户均单量': round(h1v / n, 1),
        'Q2环比pct': round(q2_chg, 1) if pd.notna(q2_chg) else None,
        'Q2同比25Q4pct': round(q2_yoy, 1) if pd.notna(q2_yoy) else None,
        '7月环比6月pct': round(m7_chg, 1) if pd.notna(m7_chg) else None,
        '7月沉默客户': silent, '前10集中度pct': round(top10_share, 1),
        'Q2保费降超30客户': int(decl_q2),
    }


rows4 = [['分群', '客户数', 'H1保费(万)', 'H1单量', '户均保费(元)', '户均单量(件)',
          'Q2环比Q1(%)', 'Q2同比25Q4(%)', '7月环比6月(%)', '7月沉默客户(户)', '前10集中度(%)', 'Q2保费降>30%(户)']]
for name, d, desc in seg_meta:
    s = seg_summary(d)
    rows4.append([name, s['客户数'], s['H1保费万'], s['H1单量'], s['户均保费'], s['户均单量'],
                  s['Q2环比pct'], s['Q2同比25Q4pct'], s['7月环比6月pct'], s['7月沉默客户'],
                  s['前10集中度pct'], s['Q2保费降超30客户']])
all_h1 = h1['保费'].sum()
share = [seg_detail_1['H1保费'].sum(), seg_detail_2['H1保费'].sum(), seg_detail_3['H1保费'].sum()]
rows4.append(['占全量保费比(%)', '', *[round(x / all_h1 * 100, 1) for x in share], '', '', '', '', '', '', ''])
for r in rows4:
    ws4.append(r)
for c in range(1, len(rows4[0]) + 1):
    cell = ws4.cell(1, c)
    cell.fill, cell.font, cell.alignment, cell.border = HEADER_FILL, HEADER_FONT, CENTER, BORDER
for i in range(2, len(rows4) + 1):
    for c in range(1, len(rows4[0]) + 1):
        ws4.cell(i, c).border = BORDER
        ws4.cell(i, c).alignment = CENTER
ws4.column_dimensions['A'].width = 16
for c in range(2, len(rows4[0]) + 1):
    ws4.column_dimensions[get_column_letter(c)].width = 15

# ---------- 分群明细 ----------
def write_seg_sheet(ws, d, start_rank, end_rank):
    cols = ['排名', '客户代码', 'H1保费', 'H1单量', 'Q1保费', 'Q1单量', 'Q2保费', 'Q2单量',
            'Q2保费环比%', 'Q2单量环比%', '25Q4保费', '25Q4单量', 'Q2同比25Q4%',
            '6月保费', '6月单量', '7月保费', '7月单量', '7月环比6月%', '7月状态']
    ws.append(cols)
    for c in range(1, len(cols) + 1):
        cell = ws.cell(1, c)
        cell.fill, cell.font, cell.alignment, cell.border = HEADER_FILL, HEADER_FONT, CENTER, BORDER
    d = d.sort_values('排名')
    for _, r in d.iterrows():
        row = [r['排名'], r['客户代码'], r['H1保费'], r['H1单量'], r['Q1保费'], r['Q1单量'],
               r['Q2保费'], r['Q2单量'], r['Q2保费环比%'], None, r['25Q4保费'], r['25Q4单量'],
               r['Q2同比25Q4%'], r['6月保费'], r['6月单量'], r['7月保费'], r['7月单量'],
               r['7月环比6月%'], r['7月状态']]
        q1v = r['Q1单量']
        row[9] = (r['Q2单量'] - q1v) / q1v * 100 if q1v > 0 else np.nan
        ws.append(row)
        i = ws.max_row
        for c in range(1, len(cols) + 1):
            cell = ws.cell(i, c)
            cell.border = BORDER
        for c in (3, 5, 7, 11, 14, 16):
            ws.cell(i, c).number_format = NUM_FMT
        for c in (4, 6, 8, 12, 15, 17):
            ws.cell(i, c).number_format = INT_FMT
        for c in (9, 10, 13, 18):
            ws.cell(i, c).number_format = PCT_FMT
            v = ws.cell(i, c).value
            if pd.notna(v):
                if v < -30:
                    ws.cell(i, c).fill, ws.cell(i, c).font = RED_FILL, RED_FONT
                elif v > 30:
                    ws.cell(i, c).fill, ws.cell(i, c).font = GREEN_FILL, GREEN_FONT
        if r['7月保费'] == 0:
            ws.cell(i, 19).fill, ws.cell(i, 19).font = RED_FILL, RED_FONT
        elif r['7月状态'] == '增长':
            ws.cell(i, 19).fill, ws.cell(i, 19).font = GREEN_FILL, GREEN_FONT
    widths = [6, 10, 12, 10, 12, 10, 12, 10, 12, 12, 12, 10, 12, 12, 10, 12, 10, 12, 10]
    for c, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = 'A2'


ws5 = wb.create_sheet('前150客户明细')
write_seg_sheet(ws5, seg_detail_1, 1, 150)
ws6 = wb.create_sheet('151-400客户明细')
write_seg_sheet(ws6, seg_detail_2, 151, 400)
ws7 = wb.create_sheet('401+客户明细')
write_seg_sheet(ws7, seg_detail_3, 401, len(seg_detail_3))

# ---------- 重点挽回名单（流失大户 + 本月沉默） ----------
ws8 = wb.create_sheet('重点挽回名单')
cols8 = ['分群', '排名', '客户代码', 'H1保费', 'Q1保费', 'Q2保费', 'Q2保费环比%', 'Q2单量环比%',
         '6月保费', '7月保费', '7月状态', '挽回优先级']
ws8.append(cols8)
for c in range(1, len(cols8) + 1):
    cell = ws8.cell(1, c)
    cell.fill, cell.font, cell.alignment, cell.border = HEADER_FILL, HEADER_FONT, CENTER, BORDER
rescue_rows = []
for name, d in [('前150', seg_detail_1), ('151-400', seg_detail_2), ('401+', seg_detail_3)]:
    d2 = d[(d['Q2保费环比%'] <= -30) | (d['7月保费'] == 0)].copy()
    for _, r in d2.sort_values('H1保费', ascending=False).iterrows():
        h1p = r['H1保费']
        prio = '高' if h1p >= 100000 else ('中' if h1p >= 30000 else '标准')
        q2v = r['Q2单量']
        q1v = r['Q1单量']
        q2v_chg = (q2v - q1v) / q1v * 100 if q1v > 0 else np.nan
        rescue_rows.append([name, r['排名'], r['客户代码'], h1p, r['Q1保费'], r['Q2保费'],
                            r['Q2保费环比%'], q2v_chg, r['6月保费'], r['7月保费'], r['7月状态'], prio])
for row in rescue_rows:
    ws8.append(row)
    i = ws8.max_row
    for c in range(1, len(cols8) + 1):
        ws8.cell(i, c).border = BORDER
    ws8.cell(i, 4).number_format = NUM_FMT
    for c in (5, 6, 9, 10):
        ws8.cell(i, c).number_format = NUM_FMT
    for c in (7, 8):
        ws8.cell(i, c).number_format = PCT_FMT
        v = ws8.cell(i, c).value
        if pd.notna(v) and v < -30:
            ws8.cell(i, c).fill, ws8.cell(i, c).font = RED_FILL, RED_FONT
    if ws8.cell(i, 11).value == '本月沉默':
        ws8.cell(i, 11).fill, ws8.cell(i, 11).font = RED_FILL, RED_FONT
    if ws8.cell(i, 12).value == '高':
        ws8.cell(i, 12).font = Font(color='9C0006', bold=True)
for c, w in zip(range(1, len(cols8) + 1), [10, 7, 10, 12, 12, 12, 12, 12, 12, 12, 10, 10]):
    ws8.column_dimensions[get_column_letter(c)].width = w
ws8.freeze_panes = 'A2'

wb.save(OUT)
print("已生成:", OUT)
print("异常表行数: T1=%d T2=%d T3=%d" % (len(T1), len(T2), len(T3)))
print("分群明细: 前150=%d, 151-400=%d, 401+=%d" % (len(seg_detail_1), len(seg_detail_2), len(seg_detail_3)))
print("重点挽回名单行数:", len(rescue_rows))
